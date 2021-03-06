import requests
from bs4 import BeautifulSoup
import time
import os 
from datetime import date
import openpyxl


class Main():
    def __init__(self, dicts, shell, title, link, price, date, dut):
        self.dict           = dicts
        self.shell          = shell
        self.title          = title
        self.link           = link
        self.price          = price 
        self.date           = date
        self.dut            = dut

    # getting all ours links
    def takeLinks(self):
        # all our links
        self.links = []
        # all our categories
        self.Category = []
        # write all our catigories
        for i in self.dict.keys():
            # write list function
            self.Category.append(i)
        # write all our links
        for i in self.dict:
            # write list function
            self.links.append(self.dict[i])


    def main(self):
        y = 0
        # Here we create new folder(day,month,hours,minuts)
        os.mkdir(str(self.dut))
        # this for open every first category page  and calculate how much pages we are have.
        for i in range(0 , len(self.links)):
            # first page link
            url = self.links[i]
            # get HTTP
            r = requests.get(url)
            # convert HTTP => HTML
            html = BeautifulSoup(r.content, "html.parser")
            # find numbers items in this category
            pag = html.find('span', attrs={
                "class": "totalcount"
            })
            # calculate how much pages we have
            spug = int(int(pag.text)/120)
            #print pages numbers
            print(spug)
            # print what link now we open and scrap
            print(self.links[i])
            # open file.txt
            # MODIFI this!!!!! We need use category name
            f = open(str(self.dut)+"/"+str(self.Category[y].replace('/', '-')) +'.txt','w', encoding="utf-8")
            

            # excel open
            self.wb = openpyxl.Workbook()
            ws = self.wb.active
            # resize
            ws.column_dimensions['B'].width = 100
            ws.column_dimensions['D'].width = 200

            t = 0
            
            y+=1
            # what items we skip
            z = 0
            
            # start scrap every page in this category
            for r in range(spug):
                # open next page
                pages = (requests.get(url+str("?s="+str(0+ z))))
                # skip 120 items
                z += 120
                # get HTML file page
                bhtml = BeautifulSoup(pages.content,"html.parser")
                # print what page we open now. 
                print(r)
                
                # selected all shell what we have in pages
                for el in bhtml.select(self.shell):
                    #take all data's
                    title = el.select(self.title)
                    link = el.select(self.title)
                    price = el.select(self.price)
                    date = el.select(self.date)
                    # excel saver
                    t +=1
                    ws['A'+str(t)]= date[0].text
                    ws['B'+str(t)]= title[0].text
                    ws['C'+str(t)]= price[0].text
                    ws['D'+str(t)]= link[0].get('href')


                    
                    # write all this in txt file
                    f.write(date[0].text)
                    f.write('---')
                    f.write(title[0].text)
                    f.write(' --- ')
                    f.write(price[0].text)
                    f.write(' --- ')
                    f.write(link[0].get('href'))
                    f.write('\n')

            # close file    
            f.close()
            self.wb.save(str(self.dut)+"/"+str(self.Category[y].replace('/', '-'))+".xlsx")